from Code import multi
# import noise_settings
import os
import random
# import stt
import time
# import tts
import copy
from openpyxl import Workbook
from pydub import AudioSegment
from Code import data_log


# from Code import email_excel

class headset_class:
    def __init__(self, selected_level, freq, noise_settings, human_speaker, noise_speaker):
        self.data_logs = ""
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.dest_filename = noise_settings + " @ " + str(freq) + " frequency.xlsx"
        self.testing_log = ""
        self.sheet.insert_cols(idx=1, amount=3)
        self.sheet['A1'] = "Test#"
        self.sheet['B1'] = "Command Name"
        self.sheet['C1'] = "Log Data"
        self.sheet.freeze_panes = "D2"
        self.num_silent = 0
        self.is_silent = False
        self.current_level = selected_level
        self.selected_level = selected_level
        self.frequency = freq
        self.noise_settings = noise_settings
        self.counter = 0
        self.human_speaker = human_speaker
        self.noise_speaker = noise_speaker
        self.command_paths = []
        for entry in os.scandir(self.selected_level):
            self.command_paths.append(entry.path)
        num_row = 2
        for row in range(self.frequency):
            for num in range(len(self.command_paths)):
                self.sheet.cell(column=1, row=num_row, value=row + 1)
                num_row += 1
        self.noise_paths = self.noise_settings

    # if not self.noise_settings:
    #    pass
    # elif self.noise_settings["Constant Sound"]:
    #     self.noise_paths = "Noise/" + self.noise_settings['Noise Type']
    # else:
    #    dir_length = len([name for name in os.listdir('Noise/' + self.noise_settings['Noise Type'])])
    #    if dir_length <= self.frequency:
    #        for entry in os.scandir('Noise/' + self.noise_settings['Noise Type']):
    #            self.noise_paths.append(entry.path)
    #   else:
    #       for entry in os.scandir('Noise/' + self.noise_settings['Noise Type']):
    #           if entry.path[0] <= self.frequency:
    #               self.noise_paths.append(entry.path)

    #   def say(self, text, speed=1):
    #       self.speaker.speak(text, speed)

    #    def listen(self):
    #        try:
    #            response = self.listener.listens()
    #        except:
    #            response = "没响应"

    #        return response.lower()

    def is_silent(self):
        pass

    #   def change_speaker_lang(self, lang='en'):
    #       self.speaker.change_lang(lang)
    #       self.speaker_lang = lang

    #  def change_listener_lang(self, lang='en'):
    #      self.listener.change_lang(lang)
    #      self.listener_lang = lang

    def get_commands_from_level(self):
        pass

    def last_page(self):
        pass

    def return_to_selected_level(self):
        while self.current_level != self.selected_level:
            # self.speaker.speak("转到上一页")
            self.check_current_level()

    def command_durations(self):
        directory = os.getcwd() + '/' + self.selected_level + '/'
        durations = 0
        for entry in os.scandir(directory):
            durations += multi.get_file_duration(entry.path) + 6
            print(durations)
        return durations

    def testing_time_estimate(self):
        return self.frequency * self.command_durations()

    def parse_data(self):
        # todo
        pass

    def get_data_log(self):
        self.data_logs = data_log.get_data()

    def check_current_level(self):
        pass

    def save_excel_log(self):
        self.wb.save(filename=self.dest_filename)


#       get data from logs, assign self.level to what level is found
if __name__ == '__main__':
    base_dir = os.getcwd()
    for frequency in range(20):
        for noise in [noise for noise in os.listdir("Noise") if noise.endswith('.wav')]:
            row_num = 2
            workbook = Workbook()
            sheet = workbook.active
            # noise_setting = noise_settings.get_noise_settings()
            headset = headset_class('全局', frequency, noise, 3, 3)
            # print("Estimated time for test is " + str(headset.testing_time_estimate))
            print(headset.counter)
            print(headset.frequency)
            volume_counter = 0
            while headset.counter < headset.frequency:
                print(headset.command_paths)
                temp_command_li = copy.copy(headset.command_paths)
                temp_noise_li = headset.noise_paths
                while temp_command_li:
                    # if not temp_noise_li: temp_noise_li = headset.noise_paths try: command, noise =
                    # temp_command_li.pop(random.randint(0, len(temp_command_li))), temp_noise_li.pop(0) except:
                    print(temp_command_li)
                    command = temp_command_li.pop(random.randint(0, len(temp_command_li)) - 1)
                    headset.sheet.cell(column=2, row=row_num, value=os.path.splitext(command)[0])
                    headset.sheet.cell(column=2, row=row_num, value=volume_counter)
                    print("noise: " + noise)
                    multi.init_multi_play(command, "Noise/" + noise, volume_counter)
                    # time.sleep(2)
                    headset.get_data_log()
                    # if not headset.is_silent:
                    #     headset.get_testing_log()
                    #     headset.sheet.cell(column=2, row=row_num, value=headset.testing_log)
                    #     headset.return_to_selected_level()
                    # else:
                    #     headset.sheet.cell(column=2, row=row_num, value="No Response")
                    #    headset.num_silent += 1
                    row_num += 1
                headset.counter += 1
                volume_counter += 1
            headset.save_excel_log()

    # email_excel.email_excel()
